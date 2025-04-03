#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
@author: Angelica Goncalves.

Module to convert Slack JSON files into curated Excel files.

Classes
-------
ExcelFormat

"""

# Import standard libraries:
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string


class ExcelFormat():
    """
    Class to edit Excel tables.

    Methods
    -------
    get_active_sheet()
        Return the active worksheet in the Excel workbook.

    get_sheet(sheet)
        Return the given Excel worksheet.

    add_sheet(name)
        Return a new Excel sheet in the current Excel workbook.

    rename_sheet(sheet, name)
        Change the name of a given Excel worksheet.

    rename_sheet(sheet, name)
        Change the name of a given Excel worksheet.

    set_cell_width(ws, column_widths):
        Set the width of each column in the Excel table.

    apply_highlight_to_row(ws, row, columns, cell_color,
                           font_size, font_bold, font_horiz_alignment):
        Format a row of the Excel table with the given conditions.

    format_highlight(ws, trigger_dict)
        Format the columns of the Excel table with the given conditions.

    set_font_color_in_column(ws, col_color)
        Set the font color on a given column.

    format_text_cells(ws, col_letter)
        Replace cell on given column with CR+LF and set alignment.

    add_border(cell, border)
        Add a border to a cell without modifying the previous setting.

    draw_vertical_line(ws, draw_vert_line)
        Draw a vertical at the right of the given column.

    set_allignment(ws, alignment_vertical)
        Align the text to the left and to the top of their cells.

    format_first_row(ws, header_row)
        Format the first row of a given Excel worksheet.

    set_filters(ws)
        Apply filters to the header row of a given Excel worksheet.

    save_changes()
        Save the file in the given directory.

    """

    def __init__(self, file_path, settings):
        # Retrieve the absolute path of the Excel workbook:
        self.file_path = file_path
        # Load the Excel workbook:
        self.wb = load_workbook(self.file_path)

        # Retrieve users inputs from settings.txt:
        self.column_widths = settings.get('column_widths')
        self.text_type_cols = settings.get('text_type_cols')
        self.header_row = settings.get('header_row')
        self.font_color_in_column = settings.get('font_color_in_column')
        self.highlights = settings.get('highlights')

    def get_active_sheet(self):
        """Return the active worksheet in the Excel workbook."""
        return self.wb.active

    def get_sheet(self, sheet):
        """
        Return the given Excel worksheet.

        Arguments
        ---------
        sheet : str
            Name of the Excel worksheet.

        """
        return self.wb[sheet]

    def add_sheet(self, name):
        """
        Return a new Excel sheet in the current Excel workbook.

        Arguments
        ---------
        name : str
            Name of the new Excel worksheet.

        """
        return self.wb.create_sheet(name)

    def rename_sheet(self, sheet, name):
        """
        Change the name of a given Excel worksheet.

        Arguments
        ---------
        sheet : str
            Name of the current Excel worksheet.

        name : str
            New name of the Excel worksheet.

        """
        sheet = self.wb[sheet]
        sheet.title = name

    def set_cell_width(self, ws, column_widths):
        """
        Set the width of each column in the Excel table.

        Arguments
        ---------
        ws : str
            Name of the Excel worksheet

        column_widths : dict
            Dictionary with format {column_name: int}

        """
        for col, width in column_widths.items():
            for cell in ws[1]:
                if col.lower() == str(cell.value).lower():
                    clmn_lttr = get_column_letter(cell.column)
                    ws.column_dimensions[clmn_lttr].width = width
                    break

    def apply_highlight_to_row(self, ws, row, columns, cell_color, font_size,
                               font_bold, font_horiz_alignment):
        """
        Format a row of the Excel table with the given conditions.

        Arguments
        ---------
        ws : str
            Name of the Excel worksheet

        row : int
            Index of the given row.

        columns : list
            List of column' letters where the highlights should be applied.

        cell_color : str
            Code of the desire color.

        font_size : int
            Size of the font.

        font_bold : bool
            Wheater the font of the highlighted columns should be bold.

        font_horiz_alignment : str
            Type of horizontal alignment of the text ("left", "right", ...).

        """
        for col in columns:
            ws[f'{col}{row}'].font = Font(size=font_size, bold=font_bold)
            ws[f'{col}{row}'].alignment = Alignment(
                horizontal=font_horiz_alignment)
            if cell_color == "No Fill":
                ws[f'{col}{row}'].fill = PatternFill()
            else:
                ws[f'{col}{row}'].fill = PatternFill(
                    start_color=cell_color, end_color=cell_color,
                    fill_type='solid')

    def format_highlight(self, ws, trigger_dict):
        """
        Format the rows of the Excel table with the given conditions.

        Arguments
        ---------
        ws: str
            Name of the Excel worksheet.

        trigger_dict : dict
            Dictionary with format: {
                "activate": True/False,
                "trigger": ["column_letter", "condition", value],
                "columns": [list of columns to highlight],
                "cell_color": "color_code",
                "font_size": #,
                "font_bold": True/False,
                "font_horiz_alignment": "alignment_type"
                }

        """
        activate = trigger_dict['activate']
        # If the highlight should be added to the Excel table:
        if activate is True:
            # Retrieve variabels from the input dictionary:
            trigger_name = trigger_dict['trigger'][0]
            trigger_condition = trigger_dict['trigger'][1]
            trigger_value = trigger_dict['trigger'][2]
            columns = trigger_dict['columns']
            cell_color = str(trigger_dict['cell_color'])
            font_size = int(trigger_dict['font_size'])
            font_bold = bool(trigger_dict['font_bold'])
            font_horiz_alignment = str(trigger_dict['font_horiz_alignment'])

            # Iterate through all the rows (except the header row):
            for i in range(2, ws.max_row + 1):
                cell_value = ws[f'{trigger_name}{i}'].value

                if trigger_condition == "==":
                    if isinstance(trigger_value, bool):
                        if cell_value == trigger_value \
                                or cell_value is trigger_value:
                            self.apply_highlight_to_row(ws,
                                                        i, columns, cell_color,
                                                        font_size, font_bold,
                                                        font_horiz_alignment)
                    else:
                        if cell_value == trigger_value:
                            self.apply_highlight_to_row(ws,
                                                        i, columns, cell_color,
                                                        font_size, font_bold,
                                                        font_horiz_alignment)

                elif trigger_condition == "!=":
                    if isinstance(trigger_value, bool):
                        if cell_value != trigger_value \
                                or cell_value is not trigger_value:
                            self.apply_highlight_to_row(ws,
                                                        i, columns, cell_color,
                                                        font_size, font_bold,
                                                        font_horiz_alignment)
                    else:
                        if cell_value != trigger_value:
                            self.apply_highlight_to_row(ws,
                                                        i, columns, cell_color,
                                                        font_size, font_bold,
                                                        font_horiz_alignment)

    def set_font_color_in_column(self, ws, col_color):
        """
        Set the font color on a given column.

        Arguments
        ---------
        ws : str
            Name of the Excel worksheet.

        col_color : list
            List with the format ["color_code", "column_letter"]

        """
        col_n = column_index_from_string(col_color[1])
        for cell in ws.iter_rows(min_col=col_n, max_col=col_n,
                                 min_row=2, max_row=ws.max_row):
            cell[0].font = Font(color=col_color[0])

    def format_text_cells(self, ws, col_letter):
        """
        Replace cell on given column with CR+LF and set alignment.

        Arguments
        ---------
        ws : str
            Name of the Excel worksheet.

        col_letter : str
            Identification letter of the column.

        """
        col_n = column_index_from_string(col_letter)
        for row in ws.iter_rows(
                min_col=col_n, max_col=col_n,
                min_row=2, max_row=ws.max_row):
            for cell in row:
                # Check if the cell contains text:
                if isinstance(cell.value, str):
                    # Replace CR (carriage return) and LF (line feed) with
                    # a space:
                    cell.value = cell.value.replace('\r\n', ' ')
                    cell.value = cell.value.replace('\r', ' ')
                    cell.value = cell.value.replace('\n\n', '\n')
                    cell.alignment = Alignment(
                        wrap_text=False, vertical="top", horizontal="left"
                        )

    def add_border(self, cell, border):
        """
        Add a border to a cell without modifying the previous setting.

        Arguments
        ---------
        cell : str
            Name of the Excel worksheet.

        border : TODO

        """
        current_border = cell.border
        cell.border = Border(
            left=border.left if border.left else current_border.left,
            right=border.right if border.right else current_border.right,
            top=border.top if border.top else current_border.top,
            bottom=border.bottom if border.bottom else current_border.bottom,
        )

    def draw_vertical_line(self, ws, draw_vert_line):
        """
        Draw a vertical at the right of the given column.

        Does not alter the previous border-setting of the cells.

        Arguments
        ---------
        ws : str
            Name of the Excel worksheet.

        draw_vert_line : dict
            Example {"columns": ["A", "B", "C"], "thickness": "medium"}

        """
        for col_letter in draw_vert_line["columns"]:
            col_n = column_index_from_string(col_letter)
            for row in ws.iter_rows(min_col=col_n, max_col=col_n,
                                    min_row=1, max_row=ws.max_row):
                for cell in row:
                    border = Border(
                        right=Side(style=draw_vert_line["thickness"]))
                    self.add_border(cell, border)

    def set_allignment(self, ws, alignment_vertical):
        """
        Align the text to the left and to the top of their cells.

        It does not alter the first row.
        Should be applied before any highlights.

        Arguments
        ---------
        ws : str
            Name of the Excel worksheet.

        alignment_vertical : str
            Type of vertical alignment of the text ("top", "bottom", ...).

        """
        for row in ws.iter_rows(min_col=1, max_col=ws.max_column,
                                min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.alignment = Alignment(vertical=alignment_vertical)

    def format_first_row(self, ws, header_row):
        """Format the first row of a given Excel worksheet."""
        # Freeze the first row (Row 1):
        ws.freeze_panes = 'A2'
        # Set the height of the first row:
        ws.row_dimensions[1].height = header_row["height"]
        # Apply the color and font formatting to the 1st row (Header row):
        for color, columns in header_row["cell_color"]:
            for col in columns:
                cell = ws.cell(row=1, column=column_index_from_string(col))
                # Set the cell color:
                cell.fill = PatternFill(
                    start_color=color, end_color=color, fill_type="solid"
                    )
                # Set the cell alignment:
                cell.alignment = Alignment(
                    wrap_text=True, vertical=header_row["alignment_vert"],
                    horizontal=header_row["alignment_horiz"]
                    )
                # Set the cell font:
                cell.font = Font(size=header_row["font_size"],
                                 bold=header_row["font_bold"])

    def set_filters(self, ws):
        """
        Apply filters to the header row of a given Excel worksheet.

        Arguments
        ---------
        ws : str
            Name of the Excel worksheet.

        """
        # Define the range for the filter (all columns and rows with data)
        max_col_letter = get_column_letter(ws.max_column)
        full_range = f"A1:{max_col_letter}{ws.max_row}"
        # Apply the filter
        ws.auto_filter.ref = full_range

    def save_changes(self):
        """Save the file in the given directory."""
        self.wb.save(self.file_path)
