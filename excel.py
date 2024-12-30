from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment  
from openpyxl.utils import get_column_letter

from inputs import key_wrd_text_show

class ExcelFormat():
    def __init__(self, file_path, curr_channel_name):
        self.file_path = file_path
        self.curr_channel_name = curr_channel_name
        self.wb = load_workbook(self.file_path)
        self.ws = self.wb.active

    def set_cell_width(self):
        w = 25
        column_widths = {
            'msg_id': 12, 'msg_date': 19, 'user': 15,'name': 19, 'display_name': 19, 'deactivated': 7, 'is_bot': 7, 
            'type':8, 'text':35, 'reply_count': 5, 'reply_users_count': 5, 'latest_reply_date': 17, 'thread_date': 17, 'parent_user_id': 25, 'URL(s)': 37,
            'projects_parsed':8,
            'project_name':w, 'working_on':w, 'progress_and_roadblocks':w,'progress':w, 'roadblocks':w, 'plans_for_following_week':w,'meetings':w
            #'project_name_1':w, 'working_on_1':w, 'progress_and_roadblocks_1':w,
            #'progress_1':w, 'roadblocks_1':w, 'plans_for_following_week_1':w,
            #'meetings_1':w, 'project_name_2':w, 'working_on_2':w,
            #'progress_and_roadblocks_2':w, 'progress_2':w, 'roadblocks_2':w,
            #'plans_for_following_week_2':w, 'meetings_2':w, 'project_name_3':w,
            #'working_on_3':w, 'progress_and_roadblocks_3':w, 'progress_3':w,
            #'roadblocks_3':w, 'plans_for_following_week_3':w, 'meetings_3':w         
        }
        for col, width in column_widths.items():
            for cell in self.ws[1]:  
                if col.lower() == str(cell.value).lower():  
                    clmn_lttr = get_column_letter(cell.column)
                    self.ws.column_dimensions[clmn_lttr].width = width 
                    break

    def set_cell_color(self):
        #IP20241121  fill_color when  -> "is_bot"==True  -> message's "type"=="thread"
        fill_bot = PatternFill(start_color="FBBF8F", end_color="FBBF8F", fill_type="solid")
        fill_thread = PatternFill(start_color="FBFB99", end_color="FBFB99", fill_type="solid")
        last_row = self.ws.max_row
        checkin_rows = ['Q','R','S','T','U','V','W']#,'X','Y','Z','AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK']
        for i in range(2, last_row + 1):
            ##-- Coloring 'is_bot' cells:
            if self.ws[f'g{i}'].value == "True" or self.ws[f'g{i}'].value == True:
                for col in ['C', 'D', 'E', 'F', 'G']+checkin_rows:
                    self.ws[f'{col}{i}'].fill = fill_bot
            ##-- Coloring 'thread' cells:
            if self.ws[f'H{i}'].value == "thread":
                for col in ['H', 'I']+checkin_rows:
                    self.ws[f'{col}{i}'].fill = fill_thread
    
    def set_font_color(self):
        ##-- Text:
        for cell in self.ws['I']:    #AG:changed from E to I
            cell.font = Font(color="0707C5")
        ##-- Parent_user_id:
        for cell in self.ws['N']:    #AG:changed from J to N
            cell.font = Font(color="c10105")
        ##-- Display_name:
        for cell in self.ws['E']:  
            cell.font = Font(color="c10105")

    
    def set_cell_allignment(self):
        ##-- Loop through each cell in the column_"E" >> 'text'  and replace CR+LF    #IP20241125
        #    also, set alignments
        for row in self.ws.iter_rows(min_col=9, max_col=9, min_row=2, max_row=self.ws.max_row):   #AG: Changed from 5 to 9
            for cell in row:
                if isinstance(cell.value, str):  # Check if the cell contains text
                    # Replace CR (carriage return) and LF (line feed) with a space
                    cell.value = cell.value.replace('\r\n', ' ').replace('\r', ' ').replace('\n\n', '\n')
                    cell.alignment = Alignment(wrap_text=False, vertical="top", horizontal="left")

        ##-- Data align-to-left  IP20241124  (excluding 1st row)
        for row in self.ws.iter_rows(min_col=10, max_col=11, min_row=2, max_row=self.ws.max_row):
            for cell in row:
                cell.alignment = Alignment(horizontal='center')   # 'left'
                if isinstance(cell.value, (int, float)):
                    cell.font = Font(size=12, bold=True)

        ##-- Data align-to-top  IP20241215  (excluding 1st row)
        for row in self.ws.iter_rows(min_col=1, max_col=35, min_row=2, max_row=self.ws.max_row):
            for cell in row:
                cell.alignment = Alignment(vertical='top')   

    
    def set_format_first_row(self):
        ##-- Freeze the first row (Row 1)
        self.ws.freeze_panes = 'A2'
        ##-- Set font size and bold for the first row
        font = Font(size=9, bold=True)
        ##-- Set the height of the first row
        self.ws.row_dimensions[1].height = 43 
        ##-- Define the RGB color
        fill = PatternFill(start_color="e7c9fb", end_color="e7c9fb", fill_type="solid")
        ##-- Apply the color, font formatting to the 1st row (Header row)
        for cell in self.ws[1]:
            cell.font = font
            cell.fill = fill
            #cell.alignment = Alignment(wrap_text=True) # Set wrap text for the cells in the first row 
            cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")

    def rename_sheet(self):
        ws_title = self.curr_channel_name 
        ws_title = ws_title[:31]
        self.ws.title = ws_title 

    def save_changes(self):
        self.wb.save(self.file_path)
        
        
    def IP_excel_adjustments(self):
        """ Excel file formatting/adjustments with  openpyxl (IP) """
        wb = load_workbook(self.file_path)
        ws = wb.active
        #
        
        #AG: ---------------------------------------------------
        #AG: Suggest to move to function set_cells_width. (Although lines below get overwritten later)
        ##-- Set the column width
        column_widths = {
            'A': 12, 'B': 19, 'C': 15,'D': 8, 'E': 35, 'F': 5, 'G': 5, 'H': 17, 'I': 17, 'J': 15, 
            'K': 19, 'L': 19, 'M': 19, 'N': 13, 'O': 25 , 'P': 7 , 'Q': 6  , 'R': 37      
        }
        ###-- Apply the column widths
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        #AG: ---------------------------------------------------
        ##AG: Suggest to move to the function set_font_color
        ##--  Apply font color to all cells in column 
        font_color = "0707C5"  
        for cell in ws['E']:   #text
            cell.font = Font(color=font_color)
        font_color = "c10105"  
        for cell in ws['J']:   #parent_user_id
            cell.font = Font(color=font_color)
        
        #AG: ---------------------------------------------------
        ##AG: Suggest to collect in function set_cell_allignment
        ##-- Loop through each cell in the column_"E" >> 'text'  and replace CR+LF    #IP20241125
        #    also, set alignments
        for row in ws.iter_rows(min_col=5, max_col=5, min_row=2, max_row=ws.max_row):
            for cell in row:
                if isinstance(cell.value, str):  # Check if the cell contains text
                    # Replace CR (carriage return) and LF (line feed) with a space
                    cell.value = cell.value.replace('\r\n', ' ').replace('\r', ' ').replace('\n\n', '\n')
                    cell.alignment = Alignment(wrap_text=False, vertical="top", horizontal="left")
    
        #AG: ---------------------------------------------------
        #AG: Suggest to move this reordering to the funtion get_all_messages. Much more direct to do it with dataframes in 1 line of code
        #IP20241120  re-order columns  
        #   
        # Specify the column to move  
        col_to_move_indx = 13    # N-of-clmn==(index)+1
        col_to_insert_indx = 4
        columns = list(ws.columns) # Get all columns
        col_to_move = columns[col_to_move_indx]
        col_to_insert = columns[col_to_insert_indx]  
        col_data = [cell.value for cell in col_to_move] # Get the data in the column to move
        ws.delete_cols(col_to_move_indx+1)  # Remove the column from its current position 
        ws.insert_cols(col_to_insert_indx)  # Insert the column at the destination position
        for row_idx, value in enumerate(col_data, start=1):
            ws.cell(row=row_idx, column=col_to_insert_indx, value=value)
    
        col_to_move_indx = 14    # N-of-clmn==(index)+1
        col_to_insert_indx = 5
        # Get all columns
        columns = list(ws.columns)
        col_to_move = columns[col_to_move_indx]
        col_to_insert = columns[col_to_insert_indx]  
        # Get the data in the column to move
        col_data = [cell.value for cell in col_to_move]
        # Remove the column from its current position 
        ws.delete_cols(col_to_move_indx+1)
        # Insert the column at the destination position
        ws.insert_cols(col_to_insert_indx)   
        for row_idx, value in enumerate(col_data, start=1):
            ws.cell(row=row_idx, column=col_to_insert_indx, value=value)

        #AG: ---------------------------------------------------
        ##AG: Suggest to move to the function set_cells_width. (Although lines below get overwritten later)
        ##AG column_widths_per_name_to_use_in_function = {
        ##AG    'msg_id': 12, 'msg_date': 19, 'user': 15,'name': 19, 'display_name': 19, 'type': 8, 'text': 35, 'reply_count': 5, 'reply_users_count': 5, 'latest_reply_date': 17, 
        ##AG    'thread_Date': 17, 'parent_user_name': 15, 'json_name': 19, 'json_mod_date': 19, 'channel_folder': 25 , 'is_bot': 7 , 'deactivated': 6, 'URL(s)': 37    
        ##AG} 
        ##-- re-Set the column width AFTER moving columns  IP20241124 (preserve in code, if further column-moving will be changed)
        column_widths = {
            'A': 12, 'B': 19, 'C': 15,'D': 19, 'E': 19, 'F': 8, 'G': 35, 'H': 5, 'I': 5, 'J': 17, 
            'K': 17, 'L': 15, 'M': 19, 'N': 19, 'O': 25 , 'P': 7 , 'Q': 6, 'R': 37    
        }
        ##-- Apply the column widths
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        #AG: ---------------------------------------------------
        #AG: Suggest to move this reordering to the funtion get_all_messages. Much more direct to do it with dataframes in 1 line of code
        # IP20241124 move "deactivated" column
        col_to_move_indx = 16    # N-of-clmn==(index)+1
        col_to_insert_indx = 6
        # Get all columns
        columns = list(ws.columns)
        col_to_move = columns[col_to_move_indx]
        col_to_insert = columns[col_to_insert_indx]  
        # Get the data in the column to move
        col_data = [cell.value for cell in col_to_move]
        # Remove the column from its current position 
        ws.delete_cols(col_to_move_indx+1)
        # Insert the column at the destination position
        ws.insert_cols(col_to_insert_indx)   
        for row_idx, value in enumerate(col_data, start=1):
            ws.cell(row=row_idx, column=col_to_insert_indx, value=value)
        
        # IP20241124 move "is_bot" column
        col_to_move_indx = 16    # N-of-clmn==(index)+1
        col_to_insert_indx = 7
        # Get all columns
        columns = list(ws.columns)
        col_to_move = columns[col_to_move_indx]
        col_to_insert = columns[col_to_insert_indx]  
        # Get the data in the column to move
        col_data = [cell.value for cell in col_to_move]
        # Remove the column from its current position 
        ws.delete_cols(col_to_move_indx+1)
        # Insert the column at the destination position
        ws.insert_cols(col_to_insert_indx)   
        for row_idx, value in enumerate(col_data, start=1):
            ws.cell(row=row_idx, column=col_to_insert_indx, value=value)
        
        #AG: ---------------------------------------------------
        #AG: Suggest to move to function set_cell_allignment
        ##-- Data align-to-left  IP20241124  (excluding 1st row)
        for row in ws.iter_rows(min_col=10, max_col=11, min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.alignment = Alignment(horizontal='center')   # 'left'
                if isinstance(cell.value, (int, float)):
                    cell.font = Font(size=12, bold=True)
     
        #AG: ---------------------------------------------------
        #AG: Suggest to move to function set_format_first_row:
        #  first row (Row 1) formattings
        ##-- Freeze the first row (Row 1)
        ws.freeze_panes = 'A2'
        ##-- Set font size and bold for the first row
        font = Font(size=9, bold=True)
        ##-- Set the height of the first row
        ws.row_dimensions[1].height = 43 
        ##-- Define the RGB color
        fill = PatternFill(start_color="e7c9fb", end_color="e7c9fb", fill_type="solid")
        ##-- Apply the color, font formatting to the 1st row (Header row)
        for cell in ws[1]:
            cell.font = font
            cell.fill = fill
            #cell.alignment = Alignment(wrap_text=True) # Set wrap text for the cells in the first row 
            cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
        
        #AG: ---------------------------------------------------
        #AG: Suggest to move to function set_font_color
        font_color = "c10105"  #IP font_color User_name
        for cell in ws['E']: 
            cell.font = Font(color=font_color)

        #AG: ---------------------------------------------------
        #AG: Suggest to move to function set_cell_color
        #IP20241121  fill_color when  -> "is_bot"==True  -> message's "type"=="thread"
        fill_bot = PatternFill(start_color="FBBF8F", end_color="FBBF8F", fill_type="solid")
        fill_thread = PatternFill(start_color="FBFB99", end_color="FBFB99", fill_type="solid")
        last_row = ws.max_row
        for i in range(2, last_row + 1):
            if ws[f'g{i}'].value == "True" or ws[f'g{i}'].value == True:
                for col in ['C', 'D', 'E', 'F', 'G']:
                    ws[f'{col}{i}'].fill = fill_bot
            if ws[f'H{i}'].value == "thread":
                for col in ['H', 'I']:
                    ws[f'{col}{i}'].fill = fill_thread
        
        #AG: ---------------------------------------------------
        #IP20241129     "weekly_report" separation
        ##--    IP20241203  set widths for "weekly-report" columns
        for col_num in range(19, 33):
            col_letter = get_column_letter(col_num)   
            ws.column_dimensions[col_letter].width = 25
        
        ##-- weekly-report keywords setting:
        keywrds_wkly_report = ["Weekly Report:","Project Name:","Working on:", "Progress and Roadblocks:", "Progress:", "Roadblocks:", 
                                        "Plans for the following week:", "Meetings:"]  
        ##-- fill for Weekly-Report's titles
        fill_wkrep_titles = PatternFill(start_color="CDB5B7", end_color="CDB5B7", fill_type="solid")
    
        ##-- Get the index/letter of the last used column 
        lst_col_index = ws.max_column              # IP20241215  !!!           refactor to last filled "title-of-column-in-1st-row"
        lst_col_lttr = get_column_letter(lst_col_index)
        ##-- set columns-titles according to the  keywrds_wkly_report
    
        wkrep_tech_title = "weekly-rep-all"  # column for tech purpose - start-&-end positions of keywords 
        ws.cell(row=1, column=lst_col_index+1).value = wkrep_tech_title
        ws.cell(row=1, column=lst_col_index+1).fill = fill_wkrep_titles
        #
        for col_idx, value in zip(range(lst_col_index + 2, lst_col_index + len(keywrds_wkly_report) + 2), keywrds_wkly_report):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = value
        
            col_letter = get_column_letter(col_idx)  # Convert column number to letter
            ws.column_dimensions[col_letter].width = 25
            ws[f'{col_letter}1'].fill = fill_wkrep_titles
    
        ##-- find column for "text" of messages
        text_to_find = 'text'
        for cell in ws[1]:  # Sheet row 1 is accessed using sheet[1]
            if text_to_find.lower() in str(cell.value).lower():  # Case-insensitive search
                clmn_lttr_text = get_column_letter(cell.column)  # Get the column letter
        #print('clmn_letter_text= ',clmn_lttr_text)
    
        #
        ##-- find keywords positions in cells of the column for "text" 
        for i in range(2, last_row + 1):
            key_wrds_text = []
            cell_value = ws[f'{clmn_lttr_text}{i}'].value
            cell_value = str(cell_value).replace("*", "")     #IP20241215  replacement all asterrisks in the "text"  
            if isinstance(cell_value, str):
                for keyword in keywrds_wkly_report:
                    position = cell_value.lower().find(keyword.lower())   # .lower()
                    if position != -1:
                        key_wrds_text.append((i, keyword, position, position + len(keyword)))
                        key_wrds_text_sorted = sorted(key_wrds_text, key = lambda x: x[2])
            if key_wrds_text:
                print ( "i= ", i, " len(key_wrds_text_sorted)= ", len(key_wrds_text_sorted))
                ##-- IP20241215 delete dumb keywords, which are just a part of some complex keyword 
                if len(key_wrds_text_sorted) > 1 :
                    for j in range(len(key_wrds_text_sorted)-1 , 0, -1):   
                        print ( "j= ", j,  key_wrds_text_sorted[j][1])
                        if key_wrds_text_sorted[j][2]  >= key_wrds_text_sorted[j-1][2] and key_wrds_text_sorted[j][2]  < key_wrds_text_sorted[j-1][3]:
                            print ( "i= ", i, "key_wrds to delete= ", key_wrds_text_sorted[j][1] )
                            del key_wrds_text_sorted[j]
                #
                key_wrds_text_sorted_str = '; '.join([f"'{match[1]}' at {match[2]}-{match[3]}" for match in key_wrds_text_sorted])
                #
                ws.cell(row=i, column=lst_col_index+1, value=key_wrds_text_sorted_str)
                print("i= ", i, "key_wrds_txt_srtd= ", key_wrds_text_sorted)
                #print("i= ", i, "key_wrds_text_sorted_str= ", key_wrds_text_sorted_str)
                #
                for j in range(0, len(key_wrds_text_sorted)):    #item in key_wrds_text_sorted:
                    for cell_1 in ws[1]:   
                        title_item = key_wrds_text_sorted[j][1]
                        if title_item.lower() in str(cell_1.value).lower():
                            item_clmn=cell_1.column
                    #for j in range(len(key_wrds_text_sorted) - 1):   
                    
                    if j + 1  < len(key_wrds_text_sorted):
                        next_item = key_wrds_text_sorted[j + 1]
                        end_position = next_item[2]
                    else: 
                        next_item = ("", "", "", "")
                        end_position = len(str(cell_value))
                    print("i= ", i, "len(key_wrds_txt_srt)= ", len(key_wrds_text_sorted), "item_j= ", j, " item_clmn= ", item_clmn, "next_item[2]= ", next_item[2], f" '{key_wrds_text_sorted[j][1]}'; {key_wrds_text_sorted[j][3]}; {end_position}")
                    #
                    if key_wrd_text_show != True:    # print or not keyword as a prefix) in tthe cell 
                        key_wrd_text = ''
                    else:
                        key_wrd_text = f"'{key_wrds_text_sorted[j][1]}' "
                    #
                    cell_keywrd_value = f"{key_wrd_text}{cell_value[key_wrds_text_sorted[j][3]: end_position]}"    #IP20241215
                    #IP20241215  Replace CR (carriage return) and LF (line feed) with a space
                    cell_keywrd_value = cell_keywrd_value.replace('\r\n', ' ').replace('\r', ' ').replace('\n\n', '\n')
                    ws.cell(row=i, column=item_clmn, value = cell_keywrd_value)            #IP20241215
                    #ws.cell(row=i, column=item_clmn, value=f"{key_wrd_text}{cell_value[key_wrds_text_sorted[j][3]: end_position]}")           
        
        
        #AG: ---------------------------------------------------
        #AG: Suggest to move this reordering to the funtion get_all_messages. Much more direct to do it with dataframes in 1 line of code
        ##-- Delete columns  ::   json_name 	json_mod_date	channel_folder    #IP20241125
        #    this columns are tech only, for development and debug, not for PMs
        clmns_to_delete = ["json_name",	"json_mod_date", "channel_folder", wkrep_tech_title]
        for clmn_to_delete in clmns_to_delete:
            for cell in ws[1]:  
                if clmn_to_delete.lower() in str(cell.value).lower():  
                    ws.delete_cols(cell.column) 
                    break
        
        #AG: ---------------------------------------------------
        #AG: Suggest to move to function set_cell_width
        ##-- re-Set the column width  IP20241205
        for col in range(15,36):
            clmn_lttr = get_column_letter(col)
            ws.column_dimensions[clmn_lttr].width = 25 
        column_widths = {
            'msg_id': 12, 'msg_date': 19, 'user': 15,'name': 19, 'display_name': 19, 'deactivated': 7, 'is_bot': 7, 
            'type':8, 'text':35, 'reply_count': 5, 'reply_users_count': 5, 'latest_reply_date': 17, 'thread_date': 17, 'parent_user_id': 25, 'URL(s)': 37      
        }
        for col, width in column_widths.items():
            for cell in ws[1]:  
                if col.lower() == str(cell.value).lower():  
                    clmn_lttr = get_column_letter(cell.column)
                    ws.column_dimensions[clmn_lttr].width = width 
                    break
        
        #AG: ---------------------------------------------------
        #AG: Suggest to move to function set_cell_allignment
        ##-- Data align-to-top  IP20241215  (excluding 1st row)
        for row in ws.iter_rows(min_col=1, max_col=35, min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.alignment = Alignment(vertical='top')   
        
        #AG: ---------------------------------------------------
        #AG: Suggest to move to function rename_sheet
        ##-- Rename the sheet
        ws_title = self.curr_channel_name 
        ws_title = ws_title[:31]
        ws.title = ws_title 

        #AG: ---------------------------------------------------
        #AG: Suggest to move to function save_changes
        ##-- Save the changes to the Excel file
        wb.save(self.file_path)    

